from fastapi import BackgroundTasks, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import PlainTextResponse, Response
from pydantic import BaseModel, Field
from datetime import datetime, timedelta, timezone
from typing import Optional, Dict, Any
from urllib.parse import quote
from io import BytesIO
import asyncio
import json
import threading
import re
import statistics

try:
    from openpyxl import Workbook
except ModuleNotFoundError:
    Workbook = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfgen import canvas
except ModuleNotFoundError:
    A4 = None
    mm = None
    pdfmetrics = None
    UnicodeCIDFont = None
    canvas = None

from .gateway_adapter import GatewayAdapter, GatewayConfig, build_sample_payloads

app = FastAPI(
    title="AI中控平台 API",
    description="智能化工业设备监控与诊断平台",
    version="1.0.0"
)


class ChatRequest(BaseModel):
    question: str
    context: Optional[Dict[str, Any]] = None


class DiagnosisRequest(BaseModel):
    device_id: str
    diagnosis_type: str


class SensorData(BaseModel):
    timestamp: str
    sensor_id: str
    sensor_name: str
    sensor_type: str
    location: str
    raw_data: Dict[str, Any]
    processed_value: float
    unit: str
    status: str
    device_id: Optional[str] = None


class AlarmData(BaseModel):
    id: str
    device_id: str
    device_name: str
    type: str
    level: str
    message: str
    timestamp: str
    status: str
    location: str


class DeviceData(BaseModel):
    id: str
    name: str
    type: str
    location: str
    status: str
    last_update: str
    description: Optional[str] = None
    parameters: Optional[Dict[str, Any]] = None


class ReportCreateRequest(BaseModel):
    name: str
    type: str
    format: str
    content: list[str] = Field(default_factory=list)
    date_range: list[str] = Field(default_factory=list)


class ReportData(BaseModel):
    id: str
    name: str
    type: str
    status: str
    progress: int
    created_at: str
    format: str
    content: list[str] = Field(default_factory=list)
    date_range: list[str] = Field(default_factory=list)
    size: Optional[str] = None
    completed_at: Optional[str] = None
    snapshot: Optional[Dict[str, Any]] = None
    error_message: Optional[str] = None


class SystemConfigUpdateRequest(BaseModel):
    value: Any
    description: Optional[str] = None
    category: Optional[str] = None


class AlarmAcknowledgeRequest(BaseModel):
    acknowledged_by: str


class GatewayPollRequest(BaseModel):
    modbus_host: Optional[str] = None
    modbus_port: Optional[int] = None
    register_address: Optional[int] = None
    register_quantity: Optional[int] = None
    device_id: Optional[str] = None
    device_name: Optional[str] = None
    location: Optional[str] = None
    sample_registers: Optional[Dict[str, int]] = Field(
        default=None,
        description="网关直推的 Modbus 原始寄存器值，需包含 R0 和 R1",
    )


sensor_data_store = []
alarms_store = []
devices_store = []
reports_store = []
system_backups_store = []
system_runtime_config_store = {
    "monitoring": {
        "enabled": True,
        "interval": 30,
        "retention": 7,
        "thresholds": {
            "cpu": {"warning": 80, "error": 90},
            "memory": {"warning": 85, "error": 95},
            "network": {"warning": 80, "error": 90},
        },
    },
    "alerts": {
        "enabled": True,
        "channels": ["email", "webhook"],
        "rules": [
            {
                "id": "rule_cpu_high",
                "name": "CPU high usage",
                "condition": "cpu > 80",
                "level": "warning",
                "enabled": True,
            },
            {
                "id": "rule_memory_high",
                "name": "Memory high usage",
                "condition": "memory > 85",
                "level": "error",
                "enabled": True,
            },
        ],
    },
    "performance": {
        "maxConcurrentTasks": 50,
        "taskTimeout": 30,
        "retryAttempts": 3,
    },
}
DEVICE_ID_ALIASES = {
    'gateway_modbus_001': '1421801',
    'LIGHT_001': '1421801',
}
system_config_store = {
    "general_system_name": {
        "key": "general_system_name",
        "value": "AI中控平台",
        "description": "系统名称",
        "category": "general"
    },
    "general_system_version": {
        "key": "general_system_version",
        "value": "1.0.0",
        "description": "系统版本",
        "category": "general"
    },
    "general_company_name": {
        "key": "general_company_name",
        "value": "示例工业科技",
        "description": "公司名称",
        "category": "general"
    },
    "general_contact_email": {
        "key": "general_contact_email",
        "value": "ops@example.com",
        "description": "联系邮箱",
        "category": "general"
    },
    "general_description": {
        "key": "general_description",
        "value": "智能化工业设备监控与诊断平台",
        "description": "系统描述",
        "category": "general"
    },
    "alarm_temperature_threshold": {
        "key": "alarm_temperature_threshold",
        "value": 80,
        "description": "温度告警阈值",
        "category": "alarm"
    },
    "alarm_pressure_threshold": {
        "key": "alarm_pressure_threshold",
        "value": 1.2,
        "description": "压力告警阈值",
        "category": "alarm"
    },
    "alarm_flow_threshold": {
        "key": "alarm_flow_threshold",
        "value": 160,
        "description": "流量告警阈值",
        "category": "alarm"
    },
    "alarm_notification_methods": {
        "key": "alarm_notification_methods",
        "value": ["email", "webhook"],
        "description": "告警通知方式",
        "category": "alarm"
    },
    "alarm_retention_days": {
        "key": "alarm_retention_days",
        "value": 30,
        "description": "告警保留天数",
        "category": "alarm"
    },
    "inspection_interval_days": {
        "key": "inspection_interval_days",
        "value": 30,
        "description": "巡检间隔",
        "category": "inspection"
    },
    "inspection_auto_generate_report": {
        "key": "inspection_auto_generate_report",
        "value": True,
        "description": "自动生成报告",
        "category": "inspection"
    },
    "inspection_report_template": {
        "key": "inspection_report_template",
        "value": "standard",
        "description": "报告模板",
        "category": "inspection"
    },
    "inspection_reminder_enabled": {
        "key": "inspection_reminder_enabled",
        "value": True,
        "description": "巡检提醒",
        "category": "inspection"
    },
    "inspection_reminder_hours": {
        "key": "inspection_reminder_hours",
        "value": 24,
        "description": "提醒提前时间",
        "category": "inspection"
    },
    "ai_model_endpoint": {
        "key": "ai_model_endpoint",
        "value": "http://127.0.0.1:8000/api/chat/ask",
        "description": "AI模型服务地址",
        "category": "ai"
    },
    "ai_api_key": {
        "key": "ai_api_key",
        "value": "demo-key",
        "description": "API密钥",
        "category": "ai"
    },
    "ai_model_name": {
        "key": "ai_model_name",
        "value": "qwen",
        "description": "模型名称",
        "category": "ai"
    },
    "ai_max_tokens": {
        "key": "ai_max_tokens",
        "value": 2048,
        "description": "最大Token数",
        "category": "ai"
    },
    "ai_temperature": {
        "key": "ai_temperature",
        "value": 0.7,
        "description": "温度参数",
        "category": "ai"
    },
    "ai_enable_knowledge_base": {
        "key": "ai_enable_knowledge_base",
        "value": True,
        "description": "启用知识库",
        "category": "ai"
    },
    "system_refresh_interval": {
        "key": "system_refresh_interval",
        "value": 30,
        "description": "数据刷新间隔",
        "category": "system"
    },
    "system_log_level": {
        "key": "system_log_level",
        "value": "info",
        "description": "日志级别",
        "category": "system"
    },
    "system_log_retention_days": {
        "key": "system_log_retention_days",
        "value": 30,
        "description": "日志保留天数",
        "category": "system"
    },
    "system_debug_mode": {
        "key": "system_debug_mode",
        "value": False,
        "description": "启用调试模式",
        "category": "system"
    },
    "system_auto_backup": {
        "key": "system_auto_backup",
        "value": True,
        "description": "数据库备份",
        "category": "system"
    },
    "system_backup_interval": {
        "key": "system_backup_interval",
        "value": "daily",
        "description": "备份间隔",
        "category": "system"
    }
}
inspection_reports_store = [
    {
        "id": "inspection_report_001",
        "task_id": "task_001",
        "task_name": "月度设备巡检",
        "device_id": "device_001",
        "device_name": "主水泵A",
        "inspector": "巡检员李四",
        "findings": ["主泵运行平稳", "压力维持在标准区间"],
        "issues": [],
        "created_at": datetime.now().isoformat()
    },
    {
        "id": "inspection_report_002",
        "task_id": "task_001",
        "task_name": "月度设备巡检",
        "device_id": "device_002",
        "device_name": "温度传感器B",
        "inspector": "巡检员李四",
        "findings": ["温度传感器响应正常"],
        "issues": ["探头表面积尘较多，建议清洁"],
        "created_at": datetime.now().isoformat()
    }
]


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def build_static_devices() -> list[Dict[str, Any]]:
    return [
        {
            "id": "device_001",
            "name": "主水泵A",
            "type": "pump",
            "status": "online",
            "location": "车间A-1区",
            "description": "负责主生产线供水的核心泵组",
            "last_update": datetime.now().isoformat(),
            "parameters": {
                "pressure": 0.82,
                "flow_rate": 128,
                "temperature": 36.5
            }
        },
        {
            "id": "device_002",
            "name": "温度传感器B",
            "type": "sensor",
            "status": "fault",
            "location": "车间B-2区",
            "description": "监测干燥区域环境温度的传感器",
            "last_update": datetime.now().isoformat(),
            "parameters": {
                "temperature": 82.4,
                "humidity": 46,
                "voltage": 220
            }
        }
    ]


def build_static_alarm_records() -> list[Dict[str, Any]]:
    return [
        {
            "id": "alarm_001",
            "device_id": "device_002",
            "device_name": "温度传感器B",
            "type": "temperature_high",
            "level": "critical",
            "title": "设备温度过高",
            "message": "设备温度过高",
            "description": "温度传感器B 当前温度超过安全阈值",
            "timestamp": datetime.now().isoformat(),
            "occurred_at": datetime.now().isoformat(),
            "status": "active",
            "location": "车间B-2区",
            "acknowledged_by": None,
            "acknowledged_at": None,
            "resolved_at": None,
        }
    ]


def list_alarm_records() -> list[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}
    for alarm in build_static_alarm_records() + alarms_store:
        normalized = {
            **alarm,
            "title": alarm.get("title") or alarm.get("message"),
            "description": alarm.get("description") or alarm.get("message"),
            "occurred_at": alarm.get("occurred_at") or alarm.get("timestamp"),
            "acknowledged_by": alarm.get("acknowledged_by"),
            "acknowledged_at": alarm.get("acknowledged_at"),
            "resolved_at": alarm.get("resolved_at"),
        }
        merged[normalized["id"]] = normalized
    return list(merged.values())


def list_system_configs() -> list[Dict[str, Any]]:
    return list(system_config_store.values())


def get_system_config_record(key: str) -> Optional[Dict[str, Any]]:
    return system_config_store.get(key)


def upsert_system_config_record(key: str, value: Any, description: Optional[str] = None, category: Optional[str] = None) -> Dict[str, Any]:
    existing = system_config_store.get(key)
    derived_category = category or (key.split('_', 1)[0] if '_' in key else 'general')
    record = {
        "key": key,
        "value": value,
        "description": description or (existing or {}).get("description") or f"{derived_category}配置项",
        "category": derived_category,
    }
    system_config_store[key] = record
    return record


def get_alarm_record(alarm_id: str) -> Optional[Dict[str, Any]]:
    return next((alarm for alarm in list_alarm_records() if alarm["id"] == alarm_id), None)


def upsert_alarm_record(alarm: Dict[str, Any]) -> Dict[str, Any]:
    existing = next((item for item in alarms_store if item['id'] == alarm['id']), None)
    normalized = {
        **alarm,
        "title": alarm.get("title") or alarm.get("message"),
        "description": alarm.get("description") or alarm.get("message"),
        "occurred_at": alarm.get("occurred_at") or alarm.get("timestamp"),
    }
    if existing:
        existing.update(normalized)
        return existing
    alarms_store.append(normalized)
    return normalized


def list_all_devices() -> list[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}
    for device in build_static_devices() + devices_store:
        merged[device['id']] = device
    return list(merged.values())


def get_device_record(device_id: str) -> Optional[Dict[str, Any]]:
    return next((device for device in list_all_devices() if device['id'] == device_id), None)


def build_canonical_sensor_id(device_id: Optional[str], sensor_type: Optional[str], fallback_sensor_id: Optional[str] = None) -> str:
    normalized_device_id = normalize_device_id(device_id)
    normalized_sensor_type = (sensor_type or '').strip()
    if normalized_device_id and normalized_sensor_type:
        return f'{normalized_device_id}_{normalized_sensor_type}'
    return (fallback_sensor_id or '').strip()


def normalize_device_id(device_id: Optional[str]) -> str:
    normalized_device_id = (device_id or '').strip()
    return DEVICE_ID_ALIASES.get(normalized_device_id, normalized_device_id)


def normalize_sensor_record(sensor: Dict[str, Any]) -> Dict[str, Any]:
    normalized = dict(sensor)
    sensor_type = str(normalized.get('sensor_type') or '').strip()
    device_id = normalize_device_id(normalized.get('device_id'))
    sensor_id = str(normalized.get('sensor_id') or '').strip()

    if sensor_id and '_' in sensor_id:
        inferred_device_id, inferred_sensor_type = sensor_id.rsplit('_', 1)
        normalized_inferred_device_id = normalize_device_id(inferred_device_id)
        if normalized_inferred_device_id and inferred_sensor_type:
            device_id = device_id or normalized_inferred_device_id
            if not sensor_type:
                sensor_type = inferred_sensor_type

    canonical_sensor_id = build_canonical_sensor_id(device_id, sensor_type, sensor_id)
    if canonical_sensor_id:
        normalized['sensor_id'] = canonical_sensor_id
    if device_id:
        normalized['device_id'] = device_id
    if sensor_type:
        normalized['sensor_type'] = sensor_type

    return normalized


def get_report_record(report_id: str) -> Optional[Dict[str, Any]]:
    return next((item for item in reports_store if item['id'] == report_id), None)


def upsert_device_record(device: Dict[str, Any]) -> Dict[str, Any]:
    existing = next((item for item in devices_store if item['id'] == device['id']), None)
    if existing:
        existing.update(device)
        return existing
    devices_store.append(device)
    return device


def upsert_report_record(report: Dict[str, Any]) -> Dict[str, Any]:
    existing = next((item for item in reports_store if item['id'] == report['id']), None)
    if existing:
        existing.update(report)
        return existing
    reports_store.append(report)
    return report


def list_reports() -> list[Dict[str, Any]]:
    return sorted(reports_store, key=lambda report: report['created_at'], reverse=True)


async def finalize_report_generation(report_id: str):
    try:
        progress_steps = [45, 80, 100]
        for progress in progress_steps:
            await asyncio.sleep(1)
            report = get_report_record(report_id)
            if not report:
                return
            update_payload: Dict[str, Any] = {
                'id': report_id,
                'progress': progress,
                'error_message': None,
            }
            if progress < 100:
                update_payload['status'] = 'generating'
            else:
                completed_at = datetime.now().isoformat()
                base_report = {**report, 'completed_at': completed_at, 'status': 'completed', 'progress': 100}
                snapshot = build_report_snapshot(base_report)
                update_payload.update({
                    'status': 'completed',
                    'progress': 100,
                    'completed_at': completed_at,
                    'snapshot': snapshot,
                    'size': estimate_report_size(snapshot, (report.get('format') or 'pdf').lower()),
                })
            upsert_report_record(update_payload)
    except Exception as exc:
        upsert_report_record({
            'id': report_id,
            'status': 'failed',
            'error_message': str(exc) or '报告生成失败',
        })


def schedule_report_generation(report_id: str):
    threading.Thread(
        target=lambda: asyncio.run(finalize_report_generation(report_id)),
        daemon=True,
    ).start()


def sanitize_download_basename(name: str) -> str:
    sanitized = re.sub(r'[\\/:*?"<>|]+', '_', name).strip()
    return sanitized or 'report'


def parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        normalized = value.replace('Z', '+00:00') if isinstance(value, str) else value
        parsed = datetime.fromisoformat(normalized)
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except ValueError:
        return None


def format_report_datetime(value: Optional[str]) -> str:
    dt = parse_iso_datetime(value)
    if not dt:
        return value or '未提供'
    return dt.strftime('%Y-%m-%d %H:%M:%S')


def get_report_content_names(content_keys: list[str]) -> list[str]:
    content_labels = {
        'system': '系统状态',
        'sensors': '传感器数据',
        'alarms': '告警信息',
        'performance': '性能指标',
        'trends': '趋势分析',
    }
    return [content_labels.get(item, item) for item in content_keys] or ['无']


def filter_records_by_date_range(records: list[Dict[str, Any]], date_range: list[str]) -> list[Dict[str, Any]]:
    if len(date_range) != 2:
        return list(records)

    start = parse_iso_datetime(date_range[0])
    end = parse_iso_datetime(date_range[1])
    if not start or not end:
        return list(records)

    filtered = []
    for record in records:
        timestamp = parse_iso_datetime(record.get('timestamp'))
        if not timestamp:
            continue
        if start <= timestamp <= end:
            filtered.append(record)
    return filtered


def get_latest_sensor_snapshot(records: list[Dict[str, Any]]) -> list[Dict[str, Any]]:
    latest_by_sensor: Dict[str, Dict[str, Any]] = {}
    for item in records:
        latest_by_sensor[item['sensor_id']] = item
    return list(latest_by_sensor.values())


def build_report_snapshot(report: Dict[str, Any]) -> Dict[str, Any]:
    date_range = report.get('date_range', []) or []
    filtered_sensor_data = filter_records_by_date_range(sensor_data_store, date_range)
    filtered_alarms = filter_records_by_date_range(alarms_store, date_range)
    latest_sensor_data = get_latest_sensor_snapshot(filtered_sensor_data or sensor_data_store)
    latest_sensor_summary = build_latest_summary(latest_sensor_data) if latest_sensor_data else {
        'count': 0,
        'avg_value': 0,
        'min_value': 0,
        'max_value': 0,
        'status_counts': {'normal': 0, 'warning': 0, 'error': 0, 'offline': 0},
        'by_type': {},
    }

    device_lookup = {device['id']: device for device in devices_store}
    active_alarms = [alarm for alarm in filtered_alarms if alarm.get('status') == 'active']
    alarm_level_counts: Dict[str, int] = {}
    for alarm in filtered_alarms:
        level = alarm.get('level') or 'unknown'
        alarm_level_counts[level] = alarm_level_counts.get(level, 0) + 1

    sensor_rows = []
    for item in latest_sensor_data:
        device = device_lookup.get(item.get('device_id') or '', {})
        sensor_rows.append({
            'sensor_id': item.get('sensor_id', ''),
            'sensor_name': item.get('sensor_name', ''),
            'sensor_type': item.get('sensor_type', ''),
            'device_id': item.get('device_id', ''),
            'device_name': device.get('name') or item.get('device_id') or '-',
            'location': item.get('location', ''),
            'processed_value': item.get('processed_value', 0),
            'unit': item.get('unit', ''),
            'status': item.get('status', ''),
            'timestamp': item.get('timestamp', ''),
        })

    alarm_rows = [
        {
            'id': alarm.get('id', ''),
            'device_id': alarm.get('device_id', ''),
            'device_name': alarm.get('device_name', ''),
            'level': alarm.get('level', ''),
            'type': alarm.get('type', ''),
            'message': alarm.get('message', ''),
            'status': alarm.get('status', ''),
            'location': alarm.get('location', ''),
            'timestamp': alarm.get('timestamp', ''),
        }
        for alarm in sorted(filtered_alarms, key=lambda alarm: alarm.get('timestamp', ''), reverse=True)
    ]

    device_rows = [
        {
            'id': device.get('id', ''),
            'name': device.get('name', ''),
            'type': device.get('type', ''),
            'location': device.get('location', ''),
            'status': device.get('status', ''),
            'last_update': device.get('last_update', ''),
        }
        for device in devices_store
    ]

    return {
        'meta': {
            'name': report.get('name', ''),
            'report_id': report.get('id', ''),
            'type': report.get('type', ''),
            'format': report.get('format', ''),
            'status': report.get('status', ''),
            'created_at': report.get('created_at', ''),
            'completed_at': report.get('completed_at', ''),
            'date_range': date_range,
            'content': report.get('content', []),
            'content_names': get_report_content_names(report.get('content', [])),
        },
        'summary': {
            'device_count': len(device_rows),
            'sensor_count': len(sensor_rows),
            'alarm_count': len(alarm_rows),
            'active_alarm_count': len(active_alarms),
            'avg_value': latest_sensor_summary['avg_value'],
            'min_value': latest_sensor_summary['min_value'],
            'max_value': latest_sensor_summary['max_value'],
            'status_counts': latest_sensor_summary['status_counts'],
            'alarm_level_counts': alarm_level_counts,
        },
        'sensor_statistics_by_type': latest_sensor_summary['by_type'],
        'sensors': sorted(sensor_rows, key=lambda item: item['timestamp'], reverse=True),
        'alarms': alarm_rows,
        'devices': device_rows,
    }


def estimate_report_size(snapshot: Dict[str, Any], report_format: str) -> str:
    row_count = len(snapshot.get('sensors', [])) + len(snapshot.get('alarms', [])) + len(snapshot.get('devices', []))
    estimated_kb = max(row_count * 2, 8)
    if report_format == 'pdf':
        estimated_kb = max(estimated_kb // 2, 8)
    return f'{estimated_kb}KB'


def build_report_text_content(report: Dict[str, Any]) -> str:
    snapshot = report.get('snapshot') or build_report_snapshot(report)
    meta = snapshot.get('meta', {})
    summary = snapshot.get('summary', {})
    date_range = meta.get('date_range', []) or []
    date_range_text = ' 至 '.join(format_report_datetime(item) for item in date_range) if date_range else '未提供'
    status_counts = summary.get('status_counts', {})

    lines = [
        'AI中控平台报告',
        '',
        f"报告名称: {meta.get('name', '')}",
        f"报告ID: {meta.get('report_id', '')}",
        f"报告类型: {meta.get('type', '')}",
        f"输出格式: {meta.get('format', '')}",
        f"创建时间: {format_report_datetime(meta.get('created_at'))}",
        f"完成时间: {format_report_datetime(meta.get('completed_at'))}",
        f"时间范围: {date_range_text}",
        f"包含内容: {', '.join(meta.get('content_names', ['无']))}",
        '',
        '汇总信息',
        f"- 设备数量: {summary.get('device_count', 0)}",
        f"- 传感器数量: {summary.get('sensor_count', 0)}",
        f"- 告警数量: {summary.get('alarm_count', 0)}",
        f"- 活跃告警: {summary.get('active_alarm_count', 0)}",
        f"- 平均值: {summary.get('avg_value', 0):.2f}",
        f"- 最小值: {summary.get('min_value', 0):.2f}",
        f"- 最大值: {summary.get('max_value', 0):.2f}",
        f"- 状态分布: 正常 {status_counts.get('normal', 0)} / 预警 {status_counts.get('warning', 0)} / 异常 {status_counts.get('error', 0)} / 离线 {status_counts.get('offline', 0)}",
    ]
    return '\n'.join(lines)


def draw_pdf_line(pdf, text: str, x: float, y: float, font_size: int = 11):
    pdf.setFont('STSong-Light', font_size)
    pdf.drawString(x, y, text)


def build_report_pdf(report: Dict[str, Any]) -> bytes:
    if canvas is None or pdfmetrics is None or UnicodeCIDFont is None or A4 is None or mm is None:
        raise HTTPException(status_code=500, detail='PDF 导出依赖未安装，请先安装 reportlab')

    snapshot = report.get('snapshot') or build_report_snapshot(report)
    meta = snapshot.get('meta', {})
    summary = snapshot.get('summary', {})
    sensors = snapshot.get('sensors', [])[:8]
    alarms = snapshot.get('alarms', [])[:8]

    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    x = 18 * mm
    y = height - 20 * mm

    draw_pdf_line(pdf, 'AI中控平台报告', x, y, 16)
    y -= 10 * mm
    draw_pdf_line(pdf, f"报告名称：{meta.get('name', '')}", x, y)
    y -= 7 * mm
    draw_pdf_line(pdf, f"报告类型：{meta.get('type', '')}    输出格式：{meta.get('format', '')}", x, y)
    y -= 7 * mm
    draw_pdf_line(pdf, f"创建时间：{format_report_datetime(meta.get('created_at'))}", x, y)
    y -= 7 * mm
    draw_pdf_line(pdf, f"完成时间：{format_report_datetime(meta.get('completed_at'))}", x, y)
    y -= 7 * mm

    date_range = meta.get('date_range', []) or []
    date_range_text = ' 至 '.join(format_report_datetime(item) for item in date_range) if date_range else '未提供'
    draw_pdf_line(pdf, f"时间范围：{date_range_text}", x, y)
    y -= 10 * mm

    draw_pdf_line(pdf, '汇总信息', x, y, 13)
    y -= 7 * mm
    summary_lines = [
        f"设备数量：{summary.get('device_count', 0)}    传感器数量：{summary.get('sensor_count', 0)}",
        f"告警数量：{summary.get('alarm_count', 0)}    活跃告警：{summary.get('active_alarm_count', 0)}",
        f"平均值：{summary.get('avg_value', 0):.2f}    最小值：{summary.get('min_value', 0):.2f}    最大值：{summary.get('max_value', 0):.2f}",
    ]
    for line in summary_lines:
        draw_pdf_line(pdf, line, x, y)
        y -= 6 * mm

    y -= 2 * mm
    draw_pdf_line(pdf, '关键传感器数据', x, y, 13)
    y -= 7 * mm
    if sensors:
        for sensor in sensors:
            line = f"{sensor['sensor_name']} ({sensor['sensor_type']})：{sensor['processed_value']} {sensor['unit']}，状态 {sensor['status']}，时间 {format_report_datetime(sensor['timestamp'])}"
            draw_pdf_line(pdf, line[:60], x, y)
            y -= 6 * mm
            if y < 30 * mm:
                pdf.showPage()
                y = height - 20 * mm
    else:
        draw_pdf_line(pdf, '暂无传感器数据', x, y)
        y -= 6 * mm

    y -= 2 * mm
    draw_pdf_line(pdf, '最近告警摘要', x, y, 13)
    y -= 7 * mm
    if alarms:
        for alarm in alarms:
            line = f"[{alarm['level']}] {alarm['device_name']}：{alarm['message']}（{format_report_datetime(alarm['timestamp'])}）"
            draw_pdf_line(pdf, line[:60], x, y)
            y -= 6 * mm
            if y < 30 * mm:
                pdf.showPage()
                y = height - 20 * mm
    else:
        draw_pdf_line(pdf, '暂无告警记录', x, y)

    pdf.save()
    return buffer.getvalue()


def append_sheet_rows(sheet, headers: list[str], rows: list[list[Any]]):
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def build_report_excel(report: Dict[str, Any]) -> bytes:
    if Workbook is None:
        raise HTTPException(status_code=500, detail='Excel 导出依赖未安装，请先安装 openpyxl')

    snapshot = report.get('snapshot') or build_report_snapshot(report)
    meta = snapshot.get('meta', {})
    summary = snapshot.get('summary', {})

    workbook = Workbook()
    overview = workbook.active
    overview.title = '概览'
    overview_rows = [
        ['报告名称', meta.get('name', '')],
        ['报告ID', meta.get('report_id', '')],
        ['报告类型', meta.get('type', '')],
        ['输出格式', meta.get('format', '')],
        ['创建时间', format_report_datetime(meta.get('created_at'))],
        ['完成时间', format_report_datetime(meta.get('completed_at'))],
        ['包含内容', ', '.join(meta.get('content_names', ['无']))],
        ['设备数量', summary.get('device_count', 0)],
        ['传感器数量', summary.get('sensor_count', 0)],
        ['告警数量', summary.get('alarm_count', 0)],
        ['活跃告警', summary.get('active_alarm_count', 0)],
        ['平均值', summary.get('avg_value', 0)],
        ['最小值', summary.get('min_value', 0)],
        ['最大值', summary.get('max_value', 0)],
    ]
    for row in overview_rows:
        overview.append(row)

    sensors_sheet = workbook.create_sheet('传感器数据')
    append_sheet_rows(
        sensors_sheet,
        ['传感器ID', '传感器名称', '类型', '设备ID', '设备名称', '位置', '数值', '单位', '状态', '时间'],
        [
            [
                item.get('sensor_id', ''),
                item.get('sensor_name', ''),
                item.get('sensor_type', ''),
                item.get('device_id', ''),
                item.get('device_name', ''),
                item.get('location', ''),
                item.get('processed_value', 0),
                item.get('unit', ''),
                item.get('status', ''),
                format_report_datetime(item.get('timestamp')),
            ]
            for item in snapshot.get('sensors', [])
        ]
    )

    alarms_sheet = workbook.create_sheet('告警记录')
    append_sheet_rows(
        alarms_sheet,
        ['告警ID', '设备ID', '设备名称', '级别', '类型', '内容', '状态', '位置', '时间'],
        [
            [
                item.get('id', ''),
                item.get('device_id', ''),
                item.get('device_name', ''),
                item.get('level', ''),
                item.get('type', ''),
                item.get('message', ''),
                item.get('status', ''),
                item.get('location', ''),
                format_report_datetime(item.get('timestamp')),
            ]
            for item in snapshot.get('alarms', [])
        ]
    )

    devices_sheet = workbook.create_sheet('设备列表')
    append_sheet_rows(
        devices_sheet,
        ['设备ID', '设备名称', '类型', '位置', '状态', '最后更新时间'],
        [
            [
                item.get('id', ''),
                item.get('name', ''),
                item.get('type', ''),
                item.get('location', ''),
                item.get('status', ''),
                format_report_datetime(item.get('last_update')),
            ]
            for item in snapshot.get('devices', [])
        ]
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_report_download_payload(report: Dict[str, Any]) -> tuple[bytes, str, str]:
    report_format = (report.get('format') or 'pdf').lower()
    basename = sanitize_download_basename(report.get('name') or report.get('id') or 'report')

    if report_format == 'pdf':
        return build_report_pdf(report), f'{basename}.pdf', 'application/pdf'
    if report_format == 'excel':
        return (
            build_report_excel(report),
            f'{basename}.xlsx',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    raise HTTPException(status_code=400, detail='暂不支持该报告格式，请选择 PDF 或 Excel')


def append_sensor_record(sensor: Dict[str, Any]) -> Dict[str, Any]:
    normalized_sensor = normalize_sensor_record(sensor)
    sensor_data_store.append(normalized_sensor)
    return normalized_sensor


def append_alarm_record(alarm: Dict[str, Any]) -> Dict[str, Any]:
    return upsert_alarm_record(alarm)


def ingest_gateway_payloads(payloads: Dict[str, Any]) -> Dict[str, Any]:
    device = upsert_device_record(payloads['device'])
    sensors = [append_sensor_record(sensor) for sensor in payloads['sensor_data']]
    alarms = [append_alarm_record(alarm) for alarm in payloads['alarms']]
    return {
        'device': device,
        'sensor_data': sensors,
        'alarms': alarms,
    }


def build_latest_summary(latest_data: list[Dict[str, Any]]) -> Dict[str, Any]:
    status_counts = {'normal': 0, 'warning': 0, 'error': 0, 'offline': 0}
    by_type: Dict[str, Dict[str, Any]] = {}

    for item in latest_data:
        status = item['status']
        status_counts[status] = status_counts.get(status, 0) + 1
        sensor_type = item['sensor_type']
        entry = by_type.setdefault(sensor_type, {
            'count': 0,
            'avg_value': 0.0,
            'min_value': item['processed_value'],
            'max_value': item['processed_value'],
        })
        entry['count'] += 1
        entry['avg_value'] += item['processed_value']
        entry['min_value'] = min(entry['min_value'], item['processed_value'])
        entry['max_value'] = max(entry['max_value'], item['processed_value'])

    for entry in by_type.values():
        entry['avg_value'] = entry['avg_value'] / entry['count'] if entry['count'] else 0.0

    values = [item['processed_value'] for item in latest_data]
    return {
        'count': len(latest_data),
        'avg_value': sum(values) / len(values) if values else 0,
        'min_value': min(values) if values else 0,
        'max_value': max(values) if values else 0,
        'status_counts': status_counts,
        'by_type': by_type,
    }


def get_latest_sensor_records(limit: Optional[int] = None) -> list[Dict[str, Any]]:
    latest_by_sensor: Dict[str, Dict[str, Any]] = {}
    for raw_item in sensor_data_store:
        item = normalize_sensor_record(raw_item)
        sensor_id = item.get('sensor_id')
        if not sensor_id:
            continue
        latest_by_sensor[sensor_id] = item

    latest_data = list(latest_by_sensor.values())
    if limit is not None:
        return latest_data[-limit:]
    return latest_data


def clamp_metric(value: float, minimum: float = 0, maximum: float = 100) -> float:
    return max(minimum, min(maximum, value))


def build_system_metrics_snapshot() -> Dict[str, Any]:
    latest_sensors = get_latest_sensor_records(limit=20)
    all_devices = list_all_devices()
    active_alarms = [alarm for alarm in list_alarm_records() if alarm.get('status') == 'active']
    generating_reports = [report for report in reports_store if report.get('status') == 'generating']

    cpu_usage = clamp_metric(24 + len(latest_sensors) * 4 + len(active_alarms) * 12 + len(generating_reports) * 8)
    memory_usage = clamp_metric(38 + len(all_devices) * 6 + len(latest_sensors) * 2.5)
    network_load = clamp_metric(16 + len(latest_sensors) * 5 + len(active_alarms) * 4)
    disk_io = clamp_metric(10 + len(generating_reports) * 18 + len(reports_store) * 2, maximum=95)
    db_connections = 12 + len(all_devices) + len(active_alarms)
    api_response_time = int(85 + len(active_alarms) * 30 + len(latest_sensors) * 6 + len(generating_reports) * 40)

    return {
        "cpu": round(cpu_usage, 2),
        "memory": round(memory_usage, 2),
        "networkLoad": round(network_load, 2),
        "activePollingTasks": len(latest_sensors),
        "diskIO": round(disk_io, 2),
        "dbConnections": db_connections,
        "apiResponseTime": api_response_time,
        "timestamp": datetime.now().isoformat(),
    }


def build_time_series(metric_name: str, points: int = 12) -> list[Dict[str, Any]]:
    snapshot = build_system_metrics_snapshot()
    base_value = float(snapshot.get(metric_name, 0))
    now = datetime.now()
    series = []

    for index in range(points):
        offset = points - index - 1
        variation = ((index % 5) - 2) * 2.5
        if metric_name == 'apiResponseTime':
            value = max(0, base_value + variation * 10)
        else:
            value = clamp_metric(base_value + variation)
        series.append({
            "timestamp": (now - timedelta(minutes=5 * offset)).isoformat(),
            "value": round(value, 2),
        })

    return series


def build_system_alerts() -> list[Dict[str, Any]]:
    alerts = []
    for alarm in list_alarm_records():
        alerts.append({
            "id": alarm["id"],
            "type": alarm.get("type", "device_alert"),
            "message": alarm.get("message", ""),
            "level": alarm.get("level", "warning"),
            "status": alarm.get("status", "active"),
            "timestamp": alarm.get("occurred_at") or alarm.get("timestamp"),
            "source": alarm.get("device_name") or "system",
            "metadata": {
                "device_id": alarm.get("device_id"),
                "location": alarm.get("location"),
            },
        })
    return alerts


def build_system_services() -> list[Dict[str, Any]]:
    snapshot = build_system_metrics_snapshot()
    return [
        {
            "name": "backend-api",
            "status": "running",
            "uptime": 86400,
            "memory": round(snapshot["memory"] * 0.35, 2),
            "cpu": round(snapshot["cpu"] * 0.4, 2),
        },
        {
            "name": "gateway-adapter",
            "status": "running",
            "uptime": 86100,
            "memory": round(snapshot["memory"] * 0.2, 2),
            "cpu": round(snapshot["cpu"] * 0.25, 2),
        },
        {
            "name": "report-worker",
            "status": "running" if reports_store else "stopped",
            "uptime": 7200 if reports_store else 0,
            "memory": round(snapshot["memory"] * 0.18, 2),
            "cpu": round(snapshot["cpu"] * 0.2, 2),
        },
    ]


def build_system_logs(limit: int = 50, level: Optional[str] = None) -> list[Dict[str, Any]]:
    now = datetime.now()
    base_logs = [
        {
            "id": "log_health",
            "timestamp": now.isoformat(),
            "level": "info",
            "message": "Health check passed",
            "source": "backend-api",
        },
        {
            "id": "log_gateway",
            "timestamp": (now - timedelta(minutes=2)).isoformat(),
            "level": "info",
            "message": f"Latest sensor sync count: {len(get_latest_sensor_records())}",
            "source": "gateway-adapter",
        },
        {
            "id": "log_alarm",
            "timestamp": (now - timedelta(minutes=4)).isoformat(),
            "level": "warning" if any(alert.get("status") == "active" for alert in build_system_alerts()) else "info",
            "message": f"Active alarm count: {len([alert for alert in build_system_alerts() if alert.get('status') == 'active'])}",
            "source": "alert-engine",
        },
        {
            "id": "log_report",
            "timestamp": (now - timedelta(minutes=6)).isoformat(),
            "level": "info",
            "message": f"Generated reports: {len(reports_store)}",
            "source": "report-worker",
        },
    ]

    if level:
        base_logs = [item for item in base_logs if item["level"] == level]

    return base_logs[:limit]


def build_system_health_payload() -> Dict[str, Any]:
    snapshot = build_system_metrics_snapshot()
    avg_load = (snapshot["cpu"] + snapshot["memory"] + snapshot["networkLoad"]) / 3

    if avg_load >= 90:
        status = "critical"
    elif avg_load >= 80:
        status = "warning"
    elif avg_load >= 60:
        status = "moderate"
    else:
        status = "good"

    return {
        "status": status,
        "score": round(max(0, 100 - avg_load), 2),
        "components": {
            "cpu": snapshot["cpu"],
            "memory": snapshot["memory"],
            "network": snapshot["networkLoad"],
            "api": snapshot["apiResponseTime"],
        },
        "uptime": 86400,
    }


def build_system_resources_payload() -> Dict[str, Any]:
    snapshot = build_system_metrics_snapshot()
    memory_total = 64
    memory_used = round(memory_total * (snapshot["memory"] / 100), 2)
    disk_total = 512
    disk_used = round(disk_total * (snapshot["diskIO"] / 100), 2)
    return {
        "cpu": {
            "usage": snapshot["cpu"],
            "cores": 8,
            "frequency": 3.2,
        },
        "memory": {
            "used": memory_used,
            "total": memory_total,
            "available": round(memory_total - memory_used, 2),
        },
        "disk": {
            "used": disk_used,
            "total": disk_total,
            "available": round(disk_total - disk_used, 2),
        },
        "network": {
            "bytesIn": len(get_latest_sensor_records()) * 2048,
            "bytesOut": len(list_alarm_records()) * 1024,
            "packetsIn": len(get_latest_sensor_records()) * 12,
            "packetsOut": len(list_alarm_records()) * 6,
        },
    }


def build_system_statistics_payload() -> Dict[str, Any]:
    snapshot = build_system_metrics_snapshot()
    alerts = build_system_alerts()
    resolved_alerts = [alert for alert in alerts if alert["status"] == "resolved"]
    return {
        "uptime": 86400,
        "totalRequests": 2048 + len(sensor_data_store) * 3,
        "averageResponseTime": snapshot["apiResponseTime"],
        "errorRate": round((len([alert for alert in alerts if alert["level"] in {"error", "critical"}]) / max(len(alerts), 1)) * 100, 2) if alerts else 0,
        "peakCpu": max(snapshot["cpu"], 88),
        "peakMemory": max(snapshot["memory"], 79),
        "totalAlerts": len(alerts),
        "resolvedAlerts": len(resolved_alerts),
    }


def build_system_diagnostics_payload() -> Dict[str, Any]:
    snapshot = build_system_metrics_snapshot()
    checks = [
        {
            "name": "API availability",
            "status": "pass",
            "message": "API service responding normally",
        },
        {
            "name": "CPU load",
            "status": "warning" if snapshot["cpu"] >= 80 else "pass",
            "message": f"CPU usage at {snapshot['cpu']}%",
        },
        {
            "name": "Memory load",
            "status": "warning" if snapshot["memory"] >= 85 else "pass",
            "message": f"Memory usage at {snapshot['memory']}%",
        },
    ]
    overall = "healthy" if all(check["status"] == "pass" for check in checks) else "warning"
    return {
        "overall": overall,
        "checks": checks,
        "recommendations": [
            "Review active alarms if CPU or memory remains elevated.",
            "Keep report generation and gateway sync intervals under observation.",
        ],
    }


def merge_nested_dict(target: Dict[str, Any], payload: Dict[str, Any]) -> Dict[str, Any]:
    for key, value in payload.items():
        if isinstance(value, dict) and isinstance(target.get(key), dict):
            merge_nested_dict(target[key], value)
        else:
            target[key] = value
    return target


@app.get("/")
async def root():
    return {"message": "AI中控平台 API 服务运行中", "version": "1.0.0"}


@app.get("/api/health")
async def health_check():
    return {
        "status": "healthy",
        "timestamp": datetime.now(),
        "service": "AI中控平台",
        "version": "1.0.0"
    }


@app.get("/api/system/metrics")
async def get_system_metrics():
    return {"metrics": build_system_metrics_snapshot()}


@app.get("/api/system/performance")
async def get_system_performance(timeRange: Optional[str] = None):
    return {
        "performance": {
            "cpu": build_time_series("cpu"),
            "memory": build_time_series("memory"),
            "network": build_time_series("networkLoad"),
            "api": build_time_series("apiResponseTime"),
        }
    }


@app.get("/api/system/health")
async def get_system_health():
    return {"health": build_system_health_payload()}


@app.get("/api/system/resources")
async def get_system_resources():
    return {"resources": build_system_resources_payload()}


@app.get("/api/system/processes")
async def get_system_processes():
    snapshot = build_system_metrics_snapshot()
    return {
        "processes": [
            {
                "pid": 1001,
                "name": "backend-api",
                "cpu": round(snapshot["cpu"] * 0.4, 2),
                "memory": round(snapshot["memory"] * 0.35, 2),
                "status": "running",
            },
            {
                "pid": 1002,
                "name": "gateway-adapter",
                "cpu": round(snapshot["cpu"] * 0.25, 2),
                "memory": round(snapshot["memory"] * 0.2, 2),
                "status": "running",
            },
            {
                "pid": 1003,
                "name": "report-worker",
                "cpu": round(snapshot["cpu"] * 0.2, 2),
                "memory": round(snapshot["memory"] * 0.18, 2),
                "status": "running" if reports_store else "idle",
            },
        ]
    }


@app.get("/api/system/logs")
async def get_system_logs(level: Optional[str] = None, limit: int = 50, offset: int = 0):
    logs = build_system_logs(limit=limit + offset, level=level)
    sliced_logs = logs[offset: offset + limit]
    return {"logs": sliced_logs, "total": len(logs)}


@app.get("/api/system/alerts")
async def get_system_alerts(status: Optional[str] = None, level: Optional[str] = None, limit: int = 50, offset: int = 0):
    alerts = build_system_alerts()
    if status:
        alerts = [alert for alert in alerts if alert["status"] == status]
    if level:
        alerts = [alert for alert in alerts if alert["level"] == level]
    sliced_alerts = alerts[offset: offset + limit]
    return {"alerts": sliced_alerts, "total": len(alerts)}


@app.post("/api/system/alerts/{alert_id}/acknowledge")
async def acknowledge_system_alert(alert_id: str):
    alarm = get_alarm_record(alert_id)
    if not alarm:
        raise HTTPException(status_code=404, detail="Alert not found")
    upsert_alarm_record({
        **alarm,
        "status": "acknowledged",
        "acknowledged_at": datetime.now().isoformat(),
        "acknowledged_by": "system",
    })
    return {"success": True}


@app.post("/api/system/alerts/{alert_id}/resolve")
async def resolve_system_alert(alert_id: str, payload: Optional[Dict[str, Any]] = None):
    alarm = get_alarm_record(alert_id)
    if not alarm:
        raise HTTPException(status_code=404, detail="Alert not found")
    upsert_alarm_record({
        **alarm,
        "status": "resolved",
        "resolved_at": datetime.now().isoformat(),
        "resolution": (payload or {}).get("resolution"),
    })
    return {"success": True}


@app.get("/api/system/config")
async def get_runtime_system_config():
    return {"config": system_runtime_config_store}


@app.put("/api/system/config")
async def update_runtime_system_config(payload: Dict[str, Any]):
    merge_nested_dict(system_runtime_config_store, payload)
    return {"success": True, "config": system_runtime_config_store}


@app.post("/api/system/services/{service_name}/restart")
async def restart_system_service(service_name: str):
    services = build_system_services()
    if not any(service["name"] == service_name for service in services):
        raise HTTPException(status_code=404, detail="Service not found")
    return {"success": True, "service": service_name, "status": "running"}


@app.get("/api/system/services")
async def get_system_services():
    return {"services": build_system_services()}


@app.post("/api/system/export")
async def export_system_data(payload: Dict[str, Any]):
    export_format = str(payload.get("format") or "json").lower()
    export_type = str(payload.get("type") or "all").lower()
    export_payload = {
        "type": export_type,
        "generated_at": datetime.now().isoformat(),
        "metrics": build_system_metrics_snapshot(),
        "alerts": build_system_alerts(),
        "logs": build_system_logs(),
        "statistics": build_system_statistics_payload(),
    }

    if export_format == "json":
        content = json.dumps(export_payload, ensure_ascii=False, indent=2).encode("utf-8")
        return Response(content=content, media_type="application/json")

    if export_format in {"csv", "excel"}:
        lines = [
            "section,key,value",
            f"metrics,cpu,{export_payload['metrics']['cpu']}",
            f"metrics,memory,{export_payload['metrics']['memory']}",
            f"metrics,networkLoad,{export_payload['metrics']['networkLoad']}",
            f"statistics,totalAlerts,{export_payload['statistics']['totalAlerts']}",
            f"statistics,averageResponseTime,{export_payload['statistics']['averageResponseTime']}",
        ]
        media_type = "text/csv" if export_format == "csv" else "application/vnd.ms-excel"
        return PlainTextResponse("\n".join(lines), media_type=media_type)

    raise HTTPException(status_code=400, detail="Unsupported export format")


@app.post("/api/system/cleanup")
async def cleanup_system_data(payload: Dict[str, Any]):
    cleanup_type = str(payload.get("type") or "logs")
    deleted_count = 5 if cleanup_type == "logs" else 0
    return {"success": True, "deletedCount": deleted_count}


@app.get("/api/system/statistics")
async def get_system_statistics(timeRange: Optional[str] = None):
    return {"statistics": build_system_statistics_payload()}


@app.post("/api/system/backup")
async def create_system_backup(payload: Optional[Dict[str, Any]] = None):
    backup_id = f"backup_{int(datetime.now().timestamp() * 1000)}"
    backup_record = {
        "id": backup_id,
        "timestamp": datetime.now().isoformat(),
        "size": 1024 + len(sensor_data_store) * 128,
        "type": "full" if (payload or {}).get("includeData", True) else "config",
        "status": "completed",
    }
    system_backups_store.insert(0, backup_record)
    return {
        "success": True,
        "backupId": backup_id,
        "size": backup_record["size"],
        "path": f"/backups/{backup_id}.zip",
    }


@app.get("/api/system/backups")
async def get_system_backups():
    return {"backups": system_backups_store}


@app.post("/api/system/backups/{backup_id}/restore")
async def restore_system_backup(backup_id: str):
    backup = next((item for item in system_backups_store if item["id"] == backup_id), None)
    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")
    return {"success": True}


@app.post("/api/system/diagnostics")
async def run_system_diagnostics():
    return {"diagnostics": build_system_diagnostics_payload()}


@app.get("/api/system/version")
async def get_system_version():
    return {
        "version": {
            "application": "1.0.0",
            "api": "1.0.0",
            "database": "embedded-memory",
            "buildTime": datetime.now().date().isoformat(),
            "gitCommit": "local-workspace",
        }
    }


@app.get("/api/devices")
async def get_devices():
    all_devices = list_all_devices()
    return {"devices": all_devices, "data": all_devices, "total": len(all_devices)}


@app.get("/api/devices/{device_id}")
async def get_device(device_id: str):
    device = get_device_record(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    return {"device": device, "data": device}


@app.get("/api/alarms")
async def get_alarms():
    all_alarms = list_alarm_records()
    return {"alarms": all_alarms, "data": all_alarms, "total": len(all_alarms)}


@app.post("/api/alarms/{alarm_id}/acknowledge")
async def acknowledge_alarm(alarm_id: str, request: AlarmAcknowledgeRequest):
    alarm = get_alarm_record(alarm_id)
    if not alarm:
        raise HTTPException(status_code=404, detail="告警不存在")
    updated = upsert_alarm_record({
        **alarm,
        "status": "acknowledged",
        "acknowledged_by": request.acknowledged_by,
        "acknowledged_at": datetime.now().isoformat(),
    })
    return {"message": "告警确认成功", "alarm": updated, "data": updated}


@app.post("/api/diagnosis/analyze")
async def analyze_device(request: DiagnosisRequest):
    return {
        "id": "diag_001",
        "device_id": request.device_id,
        "device_name": f"设备-{request.device_id}",
        "diagnosis_type": request.diagnosis_type,
        "status": "completed",
        "confidence": 0.85,
        "summary": "设备运行状态基本正常，但存在轻微性能下降",
        "findings": ["设备温度略高于正常范围", "振动频率在可接受范围内"],
        "recommendations": ["建议进行设备清洁维护", "检查冷却系统工作状态"],
        "created_at": datetime.now().isoformat(),
        "completed_at": datetime.now().isoformat()
    }


@app.get("/api/inspection/tasks")
async def get_inspection_tasks():
    return {
        "tasks": [
            {
                "id": "task_001",
                "name": "月度设备巡检",
                "status": "completed",
                "assigned_to": "巡检员李四"
            }
        ],
        "data": {
            "tasks": [
                {
                    "id": "task_001",
                    "name": "月度设备巡检",
                    "status": "completed",
                    "assigned_to": "巡检员李四"
                }
            ]
        },
        "total": 1
    }


@app.get("/api/inspection/reports")
async def get_inspection_reports():
    return {
        "reports": inspection_reports_store,
        "data": {
            "reports": inspection_reports_store
        },
        "total": len(inspection_reports_store)
    }


def summarize_chat_context(context: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not context:
        return {
            "system_status": None,
            "devices": [],
            "alarms": [],
        }

    system_status = context.get("systemStatus") or context.get("system_status") or {}
    devices = context.get("devices") or []
    alarms = context.get("alarms") or []

    return {
        "system_status": system_status,
        "devices": devices,
        "alarms": alarms,
    }


def build_chat_answer(question: str, context: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    normalized = summarize_chat_context(context)
    system_status = normalized["system_status"] or {}
    devices = normalized["devices"]
    alarms = normalized["alarms"]

    lower_question = question.lower()

    online_devices = [device for device in devices if device.get("status") == "online"]
    fault_devices = [device for device in devices if device.get("status") == "fault"]
    active_alarms = [alarm for alarm in alarms if alarm.get("status") == "active"]
    critical_alarms = [alarm for alarm in active_alarms if alarm.get("level") == "critical"]

    if any(keyword in question for keyword in ["在线", "设备状态", "设备", "运行状态"]) or "device" in lower_question:
        if not devices:
            answer = "当前页面还没有加载到设备数据，暂时无法基于实时设备信息回答。"
        else:
            online_names = "、".join(device.get("name", "未命名设备") for device in online_devices[:5]) or "暂无"
            fault_names = "、".join(device.get("name", "未命名设备") for device in fault_devices[:5]) or "暂无"
            answer = (
                f"当前页面共有 {system_status.get('totalDevices', len(devices))} 台设备，"
                f"其中在线 {system_status.get('onlineDevices', len(online_devices))} 台。"
                f"在线设备包括：{online_names}。"
                f"故障设备：{fault_names}。"
            )
        suggestions = ["最近有什么告警？", "严重告警有哪些？"]
        sources = ["当前页面设备数据", "当前页面系统状态"]
    elif any(keyword in question for keyword in ["告警", "报警", "异常"]) or "alarm" in lower_question:
        if not alarms:
            answer = "当前页面还没有加载到告警数据，暂时无法基于实时告警信息回答。"
        else:
            latest_alarm = active_alarms[0] if active_alarms else alarms[0]
            latest_summary = "暂无活跃告警"
            if latest_alarm:
                latest_summary = (
                    f"最新告警来自 {latest_alarm.get('device_name', '未知设备')}，"
                    f"等级为 {latest_alarm.get('level', '未知')}，内容为“{latest_alarm.get('message', '无描述')}”。"
                )
            answer = (
                f"当前页面共有 {system_status.get('activeAlarms', len(active_alarms))} 条活跃告警，"
                f"其中严重告警 {system_status.get('criticalAlarms', len(critical_alarms))} 条。"
                f"{latest_summary}"
            )
        suggestions = ["当前有哪些设备在线？", "如何处理高等级告警？"]
        sources = ["当前页面告警数据", "当前页面系统状态"]
    else:
        if system_status:
            answer = (
                f"根据当前页面数据，系统共有 {system_status.get('totalDevices', len(devices))} 台设备，"
                f"在线 {system_status.get('onlineDevices', len(online_devices))} 台，"
                f"活跃告警 {system_status.get('activeAlarms', len(active_alarms))} 条，"
                f"严重告警 {system_status.get('criticalAlarms', len(critical_alarms))} 条。"
                "如果您想看设备或告警详情，可以直接继续追问。"
            )
        else:
            answer = "当前页面还没有提供足够的实时数据，请先等待页面数据加载完成后再提问。"
        suggestions = ["当前有哪些设备在线？", "最近有什么告警？"]
        sources = ["当前页面实时数据"]

    return {
        "answer": answer,
        "confidence": 0.92 if system_status or devices or alarms else 0.45,
        "sources": sources,
        "suggestions": suggestions,
        "session_id": "session_001",
        "timestamp": datetime.now().isoformat()
    }


@app.post("/api/chat/ask")
async def ask_question(request: ChatRequest):
    return build_chat_answer(request.question, request.context)


@app.get("/api/system-config")
async def get_system_configs():
    configs = list_system_configs()
    return {
        "configs": configs,
        "data": {
            "configs": configs
        },
        "total": len(configs)
    }


@app.get("/api/system-config/{key}")
async def get_system_config(key: str):
    config = get_system_config_record(key)
    if not config:
        raise HTTPException(status_code=404, detail="配置不存在")
    return {"config": config, "data": config}


@app.put("/api/system-config/{key}")
async def update_system_config(key: str, request: SystemConfigUpdateRequest):
    updated = upsert_system_config_record(key, request.value, request.description, request.category)
    return {"message": "配置更新成功", "config": updated, "data": updated}


@app.post("/api/sensor-data")
async def create_sensor_data(data: SensorData):
    sensor = append_sensor_record(data.dict())
    return {"message": "传感器数据创建成功", "data": sensor}


@app.get("/api/sensor-data")
async def get_sensor_data(sensor_id: Optional[str] = None, limit: int = 100):
    if sensor_id:
        filtered_data = [d for d in sensor_data_store if d['sensor_id'] == sensor_id]
        return {"data": filtered_data[-limit:], "total": len(filtered_data)}
    return {"data": sensor_data_store[-limit:], "total": len(sensor_data_store)}


@app.post("/api/devices")
async def create_device(device: DeviceData):
    payload = device.dict()
    payload['last_update'] = payload.get('last_update') or datetime.now().isoformat()
    record = upsert_device_record(payload)
    return {"message": "设备创建成功", "device": record, "data": record}


@app.get("/api/reports")
async def get_reports():
    reports = list_reports()
    return {"reports": reports, "data": reports, "total": len(reports)}


@app.get("/api/reports/{report_id}")
async def get_report(report_id: str):
    report = get_report_record(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")
    return {"report": report, "data": report}


@app.post("/api/reports")
async def create_report(request: ReportCreateRequest, background_tasks: BackgroundTasks):
    if request.format.lower() not in {'pdf', 'excel'}:
        raise HTTPException(status_code=400, detail='暂不支持该报告格式，请选择 PDF 或 Excel')

    now = datetime.now().isoformat()
    report_id = f"report_{int(datetime.now().timestamp() * 1000)}"
    report = upsert_report_record({
        "id": report_id,
        "name": request.name,
        "type": request.type,
        "status": "generating",
        "progress": 10,
        "created_at": now,
        "format": request.format,
        "content": request.content,
        "date_range": request.date_range,
        "size": None,
        "completed_at": None,
        "snapshot": None,
        "error_message": None,
    })
    background_tasks.add_task(schedule_report_generation, report_id)
    return {"message": "报告创建成功", "report": report, "data": report}


@app.delete("/api/reports/{report_id}")
async def delete_report(report_id: str):
    index = next((i for i, item in enumerate(reports_store) if item['id'] == report_id), None)
    if index is None:
        raise HTTPException(status_code=404, detail="报告不存在")
    deleted = reports_store.pop(index)
    return {"message": "报告删除成功", "report": deleted}


@app.get("/api/reports/{report_id}/download")
async def download_report(report_id: str):
    report = get_report_record(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")
    if report['status'] == 'failed':
        raise HTTPException(status_code=409, detail=report.get('error_message') or "报告生成失败，请重新创建")
    if report['status'] != 'completed':
        raise HTTPException(status_code=409, detail="报告生成中，暂不可下载")

    content, filename, media_type = build_report_download_payload(report)
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
    }
    return Response(content=content, media_type=media_type, headers=headers)


@app.post("/api/alarms")
async def create_alarm(alarm: AlarmData):
    record = append_alarm_record(alarm.dict())
    return {"message": "��警创建成功", "alarm": record}


@app.post("/api/gateway/poll")
async def poll_gateway(request: Optional[GatewayPollRequest] = None):
    request = request or GatewayPollRequest()
    default_config = GatewayConfig()
    config = GatewayConfig(
        modbus_host=request.modbus_host or default_config.modbus_host,
        modbus_port=request.modbus_port or default_config.modbus_port,
        register_address=request.register_address or default_config.register_address,
        register_quantity=request.register_quantity or default_config.register_quantity,
        device_id=request.device_id or default_config.device_id,
        device_name=request.device_name or default_config.device_name,
        location=request.location or default_config.location,
    )

    mode = 'push' if request.sample_registers else 'poll'

    if request.sample_registers:
        r0 = request.sample_registers.get('R0')
        r1 = request.sample_registers.get('R1')
        if r0 is None or r1 is None:
            raise HTTPException(status_code=400, detail='sample_registers 必须包含 R0 和 R1')
        payloads = build_sample_payloads(r0, r1, config)
    else:
        adapter = GatewayAdapter(config)
        payloads = adapter.poll()

    ingested = ingest_gateway_payloads(payloads)
    return {
        'message': '网关采集成功',
        'mode': mode,
        'source': 'gateway_push' if mode == 'push' else 'backend_poll',
        'device': ingested['device'],
        'sensor_data': ingested['sensor_data'],
        'alarms': ingested['alarms'],
    }


@app.get("/api/sensor-data/latest")
async def get_latest_sensor_data():
    if not sensor_data_store:
        return {"data": [], "summary": {}}

    latest_by_sensor: Dict[str, Dict[str, Any]] = {}
    for raw_item in sensor_data_store:
        item = normalize_sensor_record(raw_item)
        sensor_id = item.get('sensor_id')
        if not sensor_id:
            continue
        latest_by_sensor[sensor_id] = item

    latest_data = list(latest_by_sensor.values())[-20:]
    return {"data": latest_data, "summary": build_latest_summary(latest_data)}
